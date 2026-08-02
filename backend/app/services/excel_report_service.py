import io
import uuid

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.services.student_service import StudentService


class ExcelReportService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.student_service = StudentService(db)

    def generate_student_list_excel(self) -> bytes:
        """Tüm öğrencilerin güncel listesini Excel olarak üretir."""
        from app.models.exam import Exam
        from app.models.result import Result

        students = self.student_service.get_all_students(limit=10000)

        wb = Workbook()
        ws = wb.active
        ws.title = "Ogrenciler"

        headers = ["Sıra", "Öğrenci Adı", "Sınıfı", "Son Sınav Neti", "Genel Başarı Yüzdesi"]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for i, s in enumerate(students, start=1):
            student_id_uuid = uuid.UUID(s["id"])

            # Son sınavı bul
            latest_exam = (
                self.db.query(Exam).join(Result).filter(Result.student_id == student_id_uuid).order_by(Exam.exam_date.desc()).first()
            )

            last_net = "-"
            success_percentage = "-"

            if latest_exam:
                results_latest = (
                    self.db.query(Result).filter(Result.student_id == student_id_uuid, Result.exam_id == latest_exam.id).all()
                )
                total_c = sum(r.correct for r in results_latest)
                total_w = sum(r.wrong for r in results_latest)
                last_net_val = total_c - (total_w * 0.25)
                last_net = f"{last_net_val:.2f}"

                all_results = self.db.query(Result).filter(Result.student_id == student_id_uuid).all()
                total_questions = sum(r.total_questions for r in all_results if r.measured)
                total_c_all = sum(r.correct for r in all_results if r.measured)

                if total_questions > 0:
                    pct = (total_c_all / total_questions) * 100
                    success_percentage = f"%{pct:.1f}"

            ws.append([i, s["full_name"], s["class_name"], last_net, success_percentage])

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 22

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def generate_student_detail_excel(self, student_id: uuid.UUID) -> bytes:
        """Bireysel öğrencinin tüm sonuç detaylarını Excel olarak üretir."""
        student = self.student_service.get_student_by_id(student_id)
        if not student:
            raise ValueError("Öğrenci bulunamadı")

        results = self.student_service.get_student_results(student_id)

        wb = Workbook()
        ws = wb.active
        ws.title = "Karne"

        ws.append([f"Öğrenci: {student['full_name']} - Sınıf: {student['class_name']}"])
        ws.merge_cells("A1:H1")
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.append([])  # Boş satır

        headers = ["Sınav", "Tarih", "Ders", "Konu", "Kazanım", "D/Y/B", "Net", "Başarı"]
        ws.append(headers)

        header_row_idx = 3
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

        for col_idx, text in enumerate(headers, start=1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r in results:
            if not r["measured"]:
                dyb = "Ölçülmedi"
                net_val = "Ölçülmedi"
                succ_val = "Ölçülmedi"
            else:
                dyb = f"{r['correct']}D {r['wrong']}Y {r['blank']}B"
                net_val = f"{r['net']:.2f}"
                succ_val = f"%{r['success_rate']:.1f}"

            ws.append(
                [r["exam_name"], r["exam_date"], r["subject_name"], r["topic_name"], r["outcome_description"], dyb, net_val, succ_val]
            )

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 40
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 10

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
